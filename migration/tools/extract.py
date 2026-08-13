#!/usr/bin/env python3
"""
old.xlsx -> data/*.json

A 1:1 extraction (Phase 1, see PLAN.md): field names and semantics match the sheet, so
the ported engine can be validated against TUNETEST before the tag system is redesigned.

What this does NOT copy over:
  - derived columns (Stats.L/M/N, Rotations.C:Q, Actions.A/Q, Builds.L, Config.O:W,
    Teams.A:I) -- the JS engine recomputes them
  - the 11 scratch sheets (see FINDINGS.md sec 1)

Normalisation, and why:
  - tag values are lowercased. Google Sheets MATCH is case-insensitive, so the sheet has
    ~800 tag cells disagreeing on capitalisation (atk/ATK, aero/Aero, basic/Basic) that
    work anyway. A JS `===` port would silently drop those buffs. See FINDINGS.md sec 4.1.
  - gear names are NOT lowercased: "CD ele ATK atk atk" and "CD ele atk atk atk" are
    genuinely different Mainstats rows.
  - stat names are canonicalised case-insensitively to their Config spelling.
  - Actions.Offtune is divided by 10000 (the sheet stores it pre-multiplied).

Unrecognised tags and dangling references are reported and, for tags, fatal.

Usage:  python tools/extract.py [--xlsx old.xlsx] [--out ../data] [--force]
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, OrderedDict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# Config!C:M enum columns -> emitted enum name. Nodes/Scalings/... are tag vocabularies;
# Stats/Scalers1-3 are all stat names (the split is only a dropdown-layout artefact).
ENUM_COLS = OrderedDict([
    ("scopes", "C"), ("nodes", "D"), ("scalings", "E"), ("elements", "F"),
    ("types", "G"), ("specials", "H"), ("actionIds", "I"),
])
STAT_COLS = ["J", "K", "L", "M"]

# Gear categories are only ever written as the header of each dropdown column in
# Config!O1:W1. Emitted as a real enum so the UI has something to bind to. Case is
# preserved here because Gear.Category preserves it.
CATEGORY_COLS = ["O", "P", "Q", "R", "S", "T", "U", "V", "W"]

# Actions tag columns -> (field name, enum to validate against)
ACTION_TAGS = OrderedDict([
    ("K", ("node", "nodes")),
    ("L", ("element", "elements")),
    ("M", ("scaling", "scalings")),
    ("N", ("type", "types")),
    ("O", ("special", "specials")),
    ("P", ("buffId", "actionIds")),
])

BUILD_SLOTS = OrderedDict([
    ("B", "resonator"), ("C", "weapon"), ("D", "mode"), ("E", "sequence"),
    ("F", "mainslot"), ("G", "sonata"), ("H", "twopc"), ("I", "mainstats"),
    ("J", "substats"),
])

# Offtune is stored x10000 in Actions.G; 39.2 is the cap the scratch sheets derive
# against, and tuneConst == MAX_OFFTUNE_RAW * tuneRate(level). Not in Config, so it is
# recorded here explicitly rather than left implicit.
OFFTUNE_SCALE = 10000
MAX_OFFTUNE = 39.2

# Deliberate, visible additions to the sheet's enums. 'midair' is used as a Special by 11
# Luuk actions but was never added to Config!H (FINDINGS.md sec 4.3). It is real data, not
# a typo -- no Stats row references it, so it is currently inert.
EXTRA_ENUMS = {"specials": ["midair"]}


class Report:
    def __init__(self) -> None:
        self.fatal: list[str] = []
        self.warn: list[str] = []

    def error(self, msg: str) -> None:
        self.fatal.append(msg)

    def warning(self, msg: str) -> None:
        self.warn.append(msg)

    def dump(self) -> bool:
        for m in self.warn:
            print(f"  warn   {m}")
        for m in self.fatal:
            print(f"  ERROR  {m}")
        return not self.fatal


def s(v) -> str:
    """Cell -> trimmed string. Mirrors the sheet's TRIM(CLEAN(...))."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).replace("\xa0", " ").strip()


def num(v, default=0.0):
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except ValueError:
        return default


def boolean(v) -> bool:
    if isinstance(v, bool):
        return v
    return s(v).upper() == "TRUE"


def rows(ws, first=2):
    for r in range(first, ws.max_row + 1):
        yield r


def clean_num(x: float) -> float:
    """Kill float noise from spreadsheet arithmetic (2/3 -> 0.6666666667)."""
    r = round(x, 10)
    return int(r) if isinstance(r, float) and r.is_integer() else r


def action_key(raw: str) -> str:
    """
    Canonical action id, applied to BOTH sides of the Actions <-> Rotations join so they
    cannot drift. The sheet's key is `Source & ": " & Action`, which leaves a trailing
    "': '" for the actions that have no sub-name ('Tune Break: ', 'none: ').
    """
    k = s(raw)
    return k[:-1].strip() if k.endswith(":") else k


# --------------------------------------------------------------------------- config

def extract_config(wb, rep: Report):
    ws = wb["Config"]
    enums = {}
    for name, col in ENUM_COLS.items():
        vals = [s(ws[f"{col}{r}"].value).lower() for r in rows(ws)]
        enums[name] = [v for v in vals if v]
    for name, extra in EXTRA_ENUMS.items():
        for v in extra:
            if v not in enums[name]:
                enums[name].append(v)
                rep.warning(f"config: added {v!r} to {name} -- used by Actions but "
                            f"missing from Config (FINDINGS sec 4.3)")

    enums["categories"] = [s(ws[f"{c}1"].value) for c in CATEGORY_COLS
                           if s(ws[f"{c}1"].value)]

    stat_names, seen = [], set()
    for col in STAT_COLS:
        for r in rows(ws):
            v = s(ws[f"{col}{r}"].value)
            if v and v.lower() not in seen:
                seen.add(v.lower())
                stat_names.append(v)
    enums["stats"] = stat_names

    cfg = {
        "constants": {
            "enemyLevel": num(ws["B1"].value),
            "resonatorLevel": num(ws["B2"].value),
            "defaultRes": num(ws["B3"].value),
            # dotConst/tuneConst are just the level-90 row of levels.json; kept for parity
            "dotConst": num(ws["B4"].value),
            "tuneConst": num(ws["B5"].value),
            "maxOfftune": MAX_OFFTUNE,
        },
        "enums": enums,
    }
    return cfg, {k: {v.lower() for v in vals} for k, vals in enums.items()}


def extract_levels(wb, rep: Report, cfg):
    """TUNEDOT: level -> dot base damage, tune rate. The only source for these."""
    ws = wb["TUNEDOT"]
    out = []
    for r in rows(ws, 4):
        lvl, dot, rate = ws[f"A{r}"].value, ws[f"B{r}"].value, ws[f"C{r}"].value
        if lvl is None or dot is None:
            continue
        out.append({"level": int(num(lvl)), "dot": clean_num(num(dot)),
                    "tuneRate": clean_num(num(rate))})

    # cross-check against the two hardcoded Config constants
    lvl = int(cfg["constants"]["resonatorLevel"])
    row = next((x for x in out if x["level"] == lvl), None)
    if row is None:
        rep.warning(f"levels: no row for resonatorLevel {lvl}")
    else:
        if row["dot"] != cfg["constants"]["dotConst"]:
            rep.warning(f"levels: dot({lvl})={row['dot']} != Config dotConst "
                        f"{cfg['constants']['dotConst']}")
        tune = row["tuneRate"] * MAX_OFFTUNE * OFFTUNE_SCALE
        if abs(tune - cfg["constants"]["tuneConst"]) > 1:
            rep.warning(f"levels: tuneRate({lvl})*maxOfftune={tune:.2f} != Config "
                        f"tuneConst {cfg['constants']['tuneConst']}")
    return out


# ----------------------------------------------------------------------------- data

def extract_gear(wb, rep: Report):
    ws = wb["Gear"]
    out, by_name = [], {}
    for r in rows(ws):
        name = s(ws[f"A{r}"].value)
        cat = s(ws[f"B{r}"].value)
        if not name or not cat:
            continue  # 'global' has no category; it is not selectable gear
        e = {"name": name, "category": cat}
        if s(ws[f"C{r}"].value):
            e["shorthand"] = s(ws[f"C{r}"].value)
        if ws[f"D{r}"].value is not None:
            e["cost"] = clean_num(num(ws[f"D{r}"].value))

        prev = by_name.get(name)
        if prev is None:
            by_name[name] = e
            out.append(e)
            continue
        if prev["category"] == cat:
            rep.warning(f"gear: exact duplicate {name!r} (row {r}); keeping one")
            continue
        # Same name, two categories -- e.g. 'Void Thunder 2pc' filed as both Sonata and
        # 2pc. Category only drives the UI dropdowns (buff lookup joins on name alone),
        # so prefer the category the name itself spells out.
        keep = e if cat.lower() in name.lower() else prev
        drop = prev if keep is e else e
        rep.warning(f"gear: {name!r} filed as both {prev['category']!r} and {cat!r} "
                    f"(row {r}); keeping {keep['category']!r}, dropping {drop['category']!r}")
        if keep is e:
            out[out.index(prev)] = e
            by_name[name] = e
    return out, set(by_name)


def extract_actions(wb, rep: Report, vocab):
    ws = wb["Actions"]
    out, seen = [], set()
    for r in rows(ws):
        source, action = s(ws[f"B{r}"].value), s(ws[f"C{r}"].value)
        if not source:
            continue
        aid = action_key(f"{source}: {action}")
        if aid in seen:
            # The sheet's MATCH resolves to the first hit, so later rows are already dead
            # weight in the current model. Dropping them keeps ids unique and parity intact.
            rep.warning(f"actions: duplicate id {aid!r} (row {r}) -- unreachable in the "
                        f"sheet too (MATCH takes the first); dropped")
            continue
        seen.add(aid)

        e = {"id": aid, "source": source, "action": action,
             "multiplier": clean_num(num(ws[f"D{r}"].value))}
        for col, field in [("E", "energy"), ("F", "concerto"), ("H", "forte1"),
                           ("I", "forte2"), ("J", "dots")]:
            v = ws[f"{col}{r}"].value
            if v is not None and s(v) != "":
                e[field] = clean_num(num(v))
        off = ws[f"G{r}"].value
        if off is not None and s(off) != "":
            e["offtune"] = clean_num(num(off) / OFFTUNE_SCALE)

        for col, (field, enum) in ACTION_TAGS.items():
            v = s(ws[f"{col}{r}"].value).lower()
            if not v:
                continue
            if v not in vocab[enum]:
                rep.error(f"actions row {r} ({aid!r}): {field}={v!r} not in Config {enum}")
            e[field] = v
        out.append(e)
    return out, seen


def extract_stats(wb, rep: Report, vocab, action_tag_values, canon_stat):
    ws = wb["Stats"]
    out = []
    for r in rows(ws):
        name = s(ws[f"B{r}"].value)
        stat = s(ws[f"F{r}"].value)
        if not name or not stat:
            continue
        e = {
            "name": name,                              # gear name: case preserved
            # scope is a controlled vocabulary, not a name, so it is lowercased to match
            # config.enums.scopes -- the sheet's capitalised values only worked because
            # its MATCH is case-insensitive.
            "scope": s(ws[f"C{r}"].value).lower(),
            "value": clean_num(num(ws[f"D{r}"].value)),
            "stat": canon_stat(stat, f"stats row {r}", rep),
            "burst": boolean(ws[f"G{r}"].value),
        }
        if e["scope"] not in vocab["scopes"]:
            rep.error(f"stats row {r} ({name!r}): bad scope {e['scope']!r}")

        tag = s(ws[f"E{r}"].value).lower()
        if tag:
            # Stats.Type draws on the *whole* flat tag namespace (elements, scalings,
            # types, specials and action ids), so validate against the union.
            if tag not in action_tag_values:
                rep.error(f"stats row {r} ({name!r}): type={tag!r} matches no action tag")
            e["type"] = tag

        for col, field in [("H", "scaler1"), ("I", "scaler2")]:
            v = s(ws[f"{col}{r}"].value)
            if v:
                e[field] = canon_stat(v, f"stats row {r} {field}", rep)
        for col, field in [("J", "start"), ("K", "end")]:
            v = ws[f"{col}{r}"].value
            if v is not None and s(v) != "":
                e[field] = clean_num(num(v))
        if s(ws[f"A{r}"].value):
            e["notes"] = s(ws[f"A{r}"].value)
        out.append(e)
    return out


def extract_rotations(wb, rep: Report, action_ids):
    """
    A rotation is a *contiguous block* of non-blank Rotations!B, which is what the sheet's
    own burst logic (Rotations!K) keys on. The name comes from any filled cell in column A
    of that block.

    Note: column A is only sparsely filled, and Calculator!A2 selects a rotation with
        FILTER(Rotations!B:B, TRIM(CLEAN(Rotations!A:A))=rotWord, Rotations!B:B<>"")
    -- an AND over *both* columns, so rows with a blank A never reached the engine. That
    was deliberate: all 51 such rows are 0% MV actions (Outro, some Lib) that contribute
    no damage, so skipping them was a speed optimisation. They are adopted into their
    rotation here, which changes damage by exactly nothing but keeps the resource totals
    (energy/concerto/offtune) honest.

    Burst is the on-field window: the LAST action whose name contains "Intro" through to
    the FIRST "Outro" at or after it, half-open. Recomputed here and cross-checked against
    the sheet's cached Rotations!K.
    """
    ws = wb["Rotations"]
    blocks: list[dict] = []
    cur: dict | None = None
    for r in rows(ws):
        act = action_key(ws[f"B{r}"].value)
        if not act:
            cur = None
            continue
        if cur is None:
            cur = {"names": [], "actions": [], "named": [], "burstSheet": []}
            blocks.append(cur)
        name = s(ws[f"A{r}"].value)
        if name:
            cur["names"].append(name)
        cur["actions"].append(act)
        cur["named"].append(bool(name))
        cur["burstSheet"].append(ws[f"K{r}"].value)

    out, names, adopted = [], set(), 0
    for b in blocks:
        uniq = list(dict.fromkeys(b["names"]))
        if not uniq:
            rep.warning(f"rotations: unnamed block of {len(b['actions'])} actions "
                        f"starting {b['actions'][0]!r} -- unreachable, dropped")
            continue
        if len(uniq) > 1:
            rep.error(f"rotations: one block carries several names {uniq} -- "
                      f"cannot tell where it splits")
            continue
        name = uniq[0]
        if name in names:
            rep.warning(f"rotations: {name!r} appears in more than one block; "
                        f"only the first is kept")
            continue
        names.add(name)
        acts = b["actions"]

        for a in acts:
            if a not in action_ids:
                rep.warning(f"rotations: {name!r} references unknown action {a!r}")

        intro = max((i for i, a in enumerate(acts) if "intro" in a.lower()), default=None)
        outro = None
        if intro is not None:
            outro = next((i for i, a in enumerate(acts)
                          if i >= intro and "outro" in a.lower()), None)
        burst = [intro is not None and outro is not None and intro <= i < outro
                 for i in range(len(acts))]
        cached = [boolean(v) for v in b["burstSheet"]]
        if burst != cached:
            bad = [i for i, (x, y) in enumerate(zip(burst, cached)) if x != y]
            rep.warning(f"rotations: {name!r} burst window disagrees with the sheet's "
                        f"cached Rotations!K at index {bad}")

        e = {"name": name, "actions": acts}
        if any(burst):
            e["burst"] = burst
        adopted += sum(1 for ok in b["named"] if not ok)
        out.append(e)
    if adopted:
        print(f"  note   adopted {adopted} action rows whose Rotations!A cell was blank "
              f"(all 0% MV, so damage is unchanged)")
    return out, names


def extract_builds(wb, rep: Report, gear_names, rotation_names):
    ws = wb["Builds"]
    out, seen = [], set()
    for r in rows(ws):
        name = s(ws[f"A{r}"].value)
        if not name:
            continue
        if name in seen:
            rep.warning(f"builds: duplicate name {name!r} (row {r})")
        seen.add(name)
        e = {"name": name}
        for col, field in BUILD_SLOTS.items():
            v = s(ws[f"{col}{r}"].value)
            if not v:
                continue
            if v not in gear_names:
                rep.warning(f"builds {name!r}: {field}={v!r} is not registered in Gear "
                            f"-- contributes nothing (FINDINGS sec 4.4)")
            e[field] = v
        rot = s(ws[f"K{r}"].value)
        if rot:
            if rot not in rotation_names:
                rep.warning(f"builds {name!r}: rotation {rot!r} does not exist "
                            f"-- no actions will be emitted")
            e["rotation"] = rot
        out.append(e)
    return out, seen


def extract_teams(wb, rep: Report, build_names):
    ws = wb["Teams"]
    out, seen = [], set()
    for r in rows(ws):
        name = s(ws[f"J{r}"].value)
        if not name:
            continue  # includes the whitespace-only placeholder rows
        if name in seen:
            # Calculator!CL resolves a row's buff list with VLOOKUP(team, Teams!J:P), which
            # takes the first hit -- so a repeated name makes every duplicate silently
            # inherit the first team's slot buffs.
            rep.warning(f"teams: duplicate name {name!r} (row {r}); the sheet would give "
                        f"it the first team's buffs")
        seen.add(name)
        slots = []
        for col in ("K", "L", "M"):
            v = s(ws[f"{col}{r}"].value)
            if v and v not in build_names:
                rep.error(f"teams {name!r}: slot build {v!r} does not exist")
            slots.append(v or None)
        e = {"name": name, "enabled": boolean(ws[f"A{r}"].value), "slots": slots}
        # Some of these are stale hand-typed fallbacks, not live results (FINDINGS sec 3
        # step 6). Kept only as a rough diff target; never as expected values.
        dpr = {}
        for col, field in [("B", "slot1"), ("D", "slot2"), ("E", "slot3"), ("F", "team")]:
            v = ws[f"{col}{r}"].value
            if isinstance(v, (int, float)):
                dpr[field] = clean_num(float(v))
        if dpr:
            e["sheetDpr"] = dpr
        out.append(e)
    return out


# ----------------------------------------------------------------------------- main

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="old.xlsx")
    ap.add_argument("--out", default="../data")
    ap.add_argument("--force", action="store_true",
                    help="write the JSON even if validation found errors")
    args = ap.parse_args()

    root = Path(__file__).resolve().parent.parent
    xlsx = (root / args.xlsx) if not Path(args.xlsx).is_absolute() else Path(args.xlsx)
    outdir = (root / args.out) if not Path(args.out).is_absolute() else Path(args.out)
    if not xlsx.exists():
        sys.exit(f"not found: {xlsx}")

    print(f"reading {xlsx.name} ...")
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=False)
    rep = Report()

    cfg, vocab = extract_config(wb, rep)
    canon = {n.lower(): n for n in cfg["enums"]["stats"]}

    def canon_stat(name: str, where: str, rep: Report) -> str:
        hit = canon.get(name.lower())
        if hit is None:
            rep.error(f"{where}: stat {name!r} is not in the Config stat vocabulary")
            return name
        return hit

    levels = extract_levels(wb, rep, cfg)
    gear, gear_names = extract_gear(wb, rep)
    actions, action_ids = extract_actions(wb, rep, vocab)

    # the flat tag namespace an action can actually present (Calculator!O = cols H:L,
    # i.e. element/scaling/type/special/buffId -- node is excluded, FINDINGS sec 4.2)
    tag_values = set()
    for a in actions:
        for f in ("element", "scaling", "type", "special", "buffId"):
            if a.get(f):
                tag_values.add(a[f])

    stats = extract_stats(wb, rep, vocab, tag_values, canon_stat)
    rotations, rotation_names = extract_rotations(wb, rep, action_ids)
    builds, build_names = extract_builds(wb, rep, gear_names, rotation_names)
    teams = extract_teams(wb, rep, build_names)

    files = {
        "config.json": cfg,
        "levels.json": levels,
        "gear.json": gear,
        "actions.json": actions,
        "stats.json": stats,
        "rotations.json": rotations,
        "builds.json": builds,
        "teams.json": teams,
    }

    print("\nvalidation:")
    ok = rep.dump()
    if not rep.warn and not rep.fatal:
        print("  clean")

    if not ok and not args.force:
        print(f"\n{len(rep.fatal)} error(s); nothing written. Re-run with --force to "
              f"write anyway.")
        return 1

    outdir.mkdir(parents=True, exist_ok=True)
    print(f"\nwriting {outdir}{Path().anchor or '/'}...".replace("/...", "/ ..."))
    for fn, payload in files.items():
        p = outdir / fn
        # newline="\n" explicitly: the default would translate to CRLF on Windows, and
        # tools/serve.py writes LF, so the files would churn on the first save from the UI.
        with open(p, "w", encoding="utf-8", newline="\n") as fh:
            fh.write(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
        n = len(payload) if isinstance(payload, list) else "-"
        print(f"  {fn:16} {str(n):>5} records  {p.stat().st_size:>8,} bytes")

    print("\nsummary")
    print(f"  gear        {len(gear):5}  " + ", ".join(
        f"{k} {v}" for k, v in Counter(g['category'] for g in gear).most_common(4)) + " ...")
    print(f"  actions     {len(actions):5}  {len(set(a['source'] for a in actions))} sources")
    print(f"  stats       {len(stats):5}  " + ", ".join(
        f"{k} {v}" for k, v in Counter(x['scope'] for x in stats).most_common()))
    print(f"  rotations   {len(rotations):5}  "
          f"{sum(len(r['actions']) for r in rotations)} action slots")
    print(f"  builds      {len(builds):5}")
    print(f"  teams       {len(teams):5}  "
          f"{sum(1 for t in teams if t['enabled'])} enabled")
    print(f"  levels      {len(levels):5}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
